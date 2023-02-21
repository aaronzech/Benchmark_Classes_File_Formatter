import openpyxl
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog



# Change the Kindergarten Grade code from KF or 25 to K.
def fixKindergartenGradeLevel(file):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    for row in range (1,1500): #stops at row 1500
        for col in range (6,7): # column J
            char = get_column_letter(col)
            # Format Grade Level Value to "K" format
            if ws[char + str(row)].value == "KF" or ws[char + str(row)].value == "25":
                ws[char + str(row)].value = 'K'
            if ws[char + str(row)].value =="Kindergarten" or ws[char + str(row)].value == "IEP Kindergarten" :
                ws[char + str(row)].value = 'K'
    wb.save(file)
            

# Format the columns of class sheet to Benchmark format.
def formatSheet(file):
    
    maxRow = 1500 # Updates this number if the file size changes, it will only process this many number of rows

    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    print(ws)
    #Concat Class Name with Last Name
    for row in range (1,maxRow): # Loop through rows 1 - bottom
     for col in range (3,4): #columns 1 -4
        char = get_column_letter(col)
        try:
            ws[char + str(row)].value = ws[get_column_letter(2) + str(row)].value + "-" +ws[get_column_letter(7) + str(row)].value
            print([char + str(row)].value)
        except:
            print("error")
    
    for row in range (2,maxRow): # Loop through rows 2 - bottom
     for col in range (5,6): #columns 1 -4
        char = get_column_letter(col)
        char2 = get_column_letter(6)
        
       
        if(ws[char2 + str(row)].value):
            ws[char + str(row)].value = "STUDENT"
        else:
             pass
             #print("ROW",row)
             #print(ws[char2 + str(row)].value)

    #change Column c
    ws['C1'].value = "Class's SIS Id"
    wb.save(file)

# Attach the teachers to the end of the class file.
# Looks at Classes_Template.xlsx file for the teacher reference
def pastInTeachers(file,file2):
    file = pd.read_excel(file,sheet_name='QRY801')
    file2 = pd.read_excel('Classes_Template.xlsx',sheet_name='Teachers')
    merge = pd.concat([file,file2],axis=0) 
    merge.to_excel('classes_appened.xlsx',index=False)

# Delete Extra columns in Newly appended sheet.
def deleteExtraColumns(file):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    ws.delete_cols(7)
    wb.save(file)

def convertToCSV(file):
    fileData = pd.read_excel(file,sheet_name='Sheet1')
    fileData.to_csv("classes_IMPORT.csv",index=False)

#Main Program 
if __name__ == "__main__":
    print("Starting...")

    root = tk.Tk()
    root.withdraw()

    print("Select Synergy class file")
    class_file = filedialog.askopenfilename()
    print("File:",class_file)
    teacher_file = 'Classes_Template.xlsx'

    fixKindergartenGradeLevel(class_file)
    formatSheet(class_file)
    pastInTeachers(class_file,teacher_file)
    deleteExtraColumns('classes_appened.xlsx')
    convertToCSV('classes_appened.xlsx')
    print("done")
